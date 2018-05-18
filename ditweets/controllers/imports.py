# -*- coding: utf-8 -*-
from ditweets import cache

def gsheet_twitter():
    from openpyxl import load_workbook
    from io import BytesIO
    import requests

    r =requests.get("https://docs.google.com/spreadsheets/d/1PPypAFZXc5Kd2IzA9G8sij3ScsBYFE0rROqtaDm0mN0/export?format=xlsx&id=1PPypAFZXc5Kd2IzA9G8sij3ScsBYFE0rROqtaDm0mN0")
    f = BytesIO(r.content)
    wb = load_workbook(f)

    ws = wb[wb.sheetnames[0]]
    from collections import OrderedDict
    comptes = OrderedDict()
    rows = list(ws.rows)
    defaults = {}
    params = ['tweets_rt','tweets_like','retweets_rt','retweets_like','likes_rt','likes_like','replies_rt','replies_like']

    for row in rows[3:]:
        if row[0].value:
            compte = row[0].value.strip()[1:]
            categorie = row[1].value.strip()
            statut = row[2].value.strip()
            if statut == u'Confirmé':
                defaults.update(dict((compte+'_'+v,"%d" % int(row[3+i].value or rows[2][3+i].value or 0)) for i,v in enumerate(params)))
                comptes[categorie] = comptes.get(categorie,[]) + [compte]


    cache['comptes'] = comptes
    cache['defaults'] = defaults

    return "ok"
