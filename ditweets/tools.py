from flask import Response,request,make_response

import datetime


def json_response(r):
    import json
    resp = Response(json.dumps(r))
    resp.headers['Content-Type'] = 'text/json'
    return resp


def dictToXls(data):
    from xlwt import Workbook,XFStyle
    from cStringIO import StringIO
    date_format = XFStyle()
    date_format.num_format_str = 'dd/mm/YYYY'
    stream = StringIO()
    wb = Workbook(encoding='utf8')
    from datetime import datetime,date
    for sheetname in data['sheets']:
        sheet = data['data'][sheetname]
        ws = wb.add_sheet(sheetname)
        for j,field in enumerate(sheet['fields']):
            fieldtxt = field[1] if isinstance(field,tuple) else field
            ws.row(0).write(j,fieldtxt)
        for i,row in enumerate(sheet['data']):
            for j,field in enumerate(sheet['fields']):
                fieldtxt = field[0] if isinstance(field,tuple) else field
                f = row.get(fieldtxt,None)
                if isinstance(f,date) or isinstance(f,datetime):
                    ws.row(i+1).write(j,f,date_format)
                else:
                    ws.row(i+1).write(j,f)

    wb.save(stream)
    return stream.getvalue()

def dictToXlsx(data):
    from openpyxl import Workbook
    from openpyxl.writer.excel import save_virtual_workbook


    wb = Workbook()
    for sheetname in data['sheets']:
        sheet = data['data'][sheetname]
        ws = wb.create_sheet(title=sheetname)
        for j,field in enumerate(sheet['fields']):
            fieldtxt = field[1] if isinstance(field,tuple) else field
            ws.cell(column=j+1,row=1,value=fieldtxt)
        for i,row in enumerate(sheet['data']):
            for j,field in enumerate(sheet['fields']):
                fieldv = field[0] if isinstance(field,tuple) else field
                f = row.get(fieldv,None)
                ws.cell(column=j+1,row=2+i,value=f)

    return save_virtual_workbook(wb)

def xls_response(filename,v):
    import datetime
    output = make_response(v)
    output.headers['Content-Disposition'] = "attachment; filename=%s_%s.xls" % (filename,datetime.datetime.now().strftime('%Y-%m-%d'))
    output.headers['Content-type'] = 'application/vnd.ms-excel'
    return output

def image_response(type,v,filename=None,nocache=True):
    headers = {'Cache-Control':'no-cache, no-store, must-revalidate','Pragma':'no-cache'} if nocache else {}
    if filename:
        headers.update({"Content-Disposition":
                     "attachment;filename=%s-%s.png" % (filename,datetime.datetime.now().strftime('%Y-%m-%d'))})
    r = Response(v, mimetype="image/%s" % type,headers=headers)
    return r

import pyzmail
from ditweets.config_private import smtp
NOTIFY_ADDRESS = ['observatoireapi@yahoo.com']
SMTP_HOST = smtp['host']

def sendmail(sender,recipients,subject,msg='',attach=[]):
    payload, mail_from, rcpt_to, msg_id=pyzmail.compose_mail(\
        sender, \
        recipients, \
        subject, \
        'utf-8', \
        (msg, 'utf-8'), \
        html=None, \
        attachments=attach)

    #[('attached content', 'text', 'plain', 'text.txt', 'utf-8')]
    smtp_host = SMTP_HOST
    ret=pyzmail.send_mail(payload, mail_from, rcpt_to, smtp_host, smtp_port=smtp['port'], smtp_mode=smtp['mode'],smtp_login=smtp['username'],smtp_password=smtp['password'])

    if isinstance(ret, dict):
        if ret:
            print('failed recipients:', ', '.join(ret.keys()))
    else:
        print('error:', ret)

def sendmail2(sender, recipient, subject, msg=''):

    import smtplib
    from email.mime.text import MIMEText

    title = subject
    msg_content = """<html><head></head><body>"""+msg+"</body></html>"
    message = MIMEText(msg_content, 'html')

    message['From'] = sender
    message['To'] = recipient
    message['Subject'] = subject

    msg_full = message.as_string()

    server = smtplib.SMTP(smtp['host'],smtp['port'])
    if smtp['mode']!='normal':
        server.starttls()
    if smtp['username']:
        server.login(smtp['username'],smtp['password'])
    server.sendmail(sender, recipient, msg_full)
    server.quit()

def api_notify(subject,msg="",attach=[],recipients=NOTIFY_ADDRESS):
    sendmail(('ditweets','api@observatoire-democratie.fr'),recipients,subject,msg,attach)
